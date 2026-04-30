"""
reporte_inprg_mc.py
-------------------
Genera un reporte Excel con las OTs MC en estado INPRG de O_GESFO.

PROPOSITO:
    Listar las OTs que estan abiertas en INPRG para revision administrativa.
    Las OTs muy antiguas (zombies) no deberian estar en este estado y deben
    ser cerradas formalmente.

ORIGEN DE DATOS:
    Postgres (schema onms). Los datos provienen del ultimo ETL.
    Si necesitas datos al minuto, ejecuta primero bandeja_o_gesfo_operativa
    para refrescar las OPERATIVAS, o el bandeja_o_gesfo_completo para todo.

ESTRUCTURA DEL EXCEL:
    Hoja 1 (OTs INPRG):  Lista detallada ordenada por antiguedad descendente
    Hoja 2 (Resumen):    Conteos por antiguedad, operador y EECC

EJECUCION:
    py -m etl.reporte_inprg_mc

SALIDA:
    output/Reporte_INPRG_MC_{YYYYMMDD_HHMMSS}.xlsx
"""

import os
import logging
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from core.logging_setup import logger
from integrations.postgres.client import obtener_conexion, cerrar_conexion


# ════════════════════════════════════════════════════════════════════
# QUERY PRINCIPAL
# ════════════════════════════════════════════════════════════════════

QUERY_INPRG = """
SELECT 
    wo.wonum,
    wo.cinum,
    wo.ci_description,
    wo.description AS resumen,
    wo.clasificacion_operativa,
    EXTRACT(EPOCH FROM (NOW() - wo.creation_date)) / 86400 AS dias_abierta,
    wo.creation_date AS fecha_creacion,
    wo.actual_start AS fecha_inicio,
    wo.assigned_to AS tecnico,
    wo.coordinador_red_fo AS coordinador,
    wo.lider_de_zona_fo AS lider_zona,
    wo.location AS codigo_sitio,
    wo.nom_ubicacion AS nombre_sitio,
    wo.direccion,
    wo.tipo_tramo,
    wo.tipo_operacion_fo,
    wo.operador_fo AS operador,
    wo.eecc_cuadrilla_fo AS eecc,
    wo.tipo_cuadrilla_fo,
    wo.numero_caso_fo AS numero_caso,
    wo.outage_asociado,
    wo.area_que_reporta_fo AS area_reporta,
    wo.persona_que_reporta,
    wo.cant_worklogs,
    ultimo.createdate AS ultimo_avance_fecha,
    ultimo.createby AS ultimo_avance_quien,
    ultimo.description AS ultimo_avance_resumen,
    ultimo.description_long AS ultimo_avance_completo,
    EXTRACT(EPOCH FROM (NOW() - ultimo.createdate)) / 86400 AS dias_sin_avance
FROM onms.work_orders wo
LEFT JOIN LATERAL (
    SELECT createdate, createby, description, description_long
    FROM onms.worklogs
    WHERE wonum = wo.wonum
    ORDER BY createdate DESC
    LIMIT 1
) ultimo ON true
WHERE wo.activa = true
  AND wo.status = 'INPRG'
  AND wo.worktype = 'MC'
ORDER BY wo.creation_date ASC
"""


# Definicion de columnas de la hoja principal
COLUMNAS_DETALLE = [
    ("wonum",                    "WONUM",                    14),
    ("dias_abierta",             "Dias Abierta",             12),
    ("dias_sin_avance",          "Dias Sin Avance",          12),
    ("clasificacion_operativa",  "Clasificacion",            18),
    ("fecha_creacion",           "Fecha Creacion",           18),
    ("resumen",                  "Resumen",                  50),
    ("tecnico",                  "Tecnico Asignado",         20),
    ("coordinador",              "Coordinador Red FO",       22),
    ("lider_zona",               "Lider Zona",               18),
    ("codigo_sitio",             "Cod. Sitio",               12),
    ("nombre_sitio",             "Nombre Sitio",             30),
    ("direccion",                "Direccion",                30),
    ("tipo_tramo",               "Tipo Tramo",               12),
    ("tipo_operacion_fo",        "Tipo Operacion",           14),
    ("operador",                 "Operador",                 22),
    ("eecc",                     "EECC Cuadrilla",           18),
    ("tipo_cuadrilla_fo",        "Tipo Cuadrilla",           20),
    ("numero_caso",              "Numero Caso",              16),
    ("outage_asociado",          "Outage",                   12),
    ("area_reporta",             "Area Reporta",             14),
    ("persona_que_reporta",      "Persona Reporta",          18),
    ("cinum",                    "CI",                       30),
    ("ci_description",           "CI Descripcion",           40),
    ("cant_worklogs",            "# Worklogs",               10),
    ("ultimo_avance_fecha",      "Ultimo Avance Fecha",      18),
    ("ultimo_avance_quien",      "Ultimo Avance Quien",      18),
    ("ultimo_avance_resumen",    "Ultimo Avance Resumen",    50),
    ("ultimo_avance_completo",   "Ultimo Avance Completo",   80),
]


# ════════════════════════════════════════════════════════════════════
# ESTILOS DE EXCEL
# ════════════════════════════════════════════════════════════════════

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

CELL_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

# Colores por antiguedad (para resaltar fila completa)
FILL_CRITICO = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")  # rosa palido
FILL_ALTO    = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")  # amarillo palido
FILL_MEDIO   = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # amarillo muy palido
FILL_BAJO    = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # verde palido


def color_por_antiguedad(dias):
    """Devuelve el PatternFill segun los dias abierta."""
    if dias is None:
        return None
    if dias > 90:
        return FILL_CRITICO
    if dias > 30:
        return FILL_ALTO
    if dias > 7:
        return FILL_MEDIO
    return FILL_BAJO


# ════════════════════════════════════════════════════════════════════
# HOJA 1: LISTA DETALLADA
# ════════════════════════════════════════════════════════════════════

def construir_hoja_detalle(ws, filas):
    """Llena la hoja con la lista detallada de OTs."""
    # Cabecera
    for col_idx, (_, header, ancho) in enumerate(COLUMNAS_DETALLE, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho

    ws.row_dimensions[1].height = 35
    ws.freeze_panes = "B2"  # Congelar primera fila y primera columna

    # Datos
    for row_idx, fila in enumerate(filas, start=2):
        dias_abierta = fila.get("dias_abierta")
        fill_fila = color_por_antiguedad(dias_abierta)

        for col_idx, (key, _, _) in enumerate(COLUMNAS_DETALLE, start=1):
            valor = fila.get(key)

            # Conversiones de formato
            if key == "dias_abierta" or key == "dias_sin_avance":
                if valor is not None:
                    valor = round(float(valor), 1)
            elif key in ("fecha_creacion", "fecha_inicio", "ultimo_avance_fecha"):
                if valor is not None:
                    valor = valor.strftime("%Y-%m-%d %H:%M") if hasattr(valor, "strftime") else str(valor)

            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.alignment = CELL_ALIGN
            cell.border = THIN_BORDER
            if fill_fila:
                cell.fill = fill_fila

    # Auto-filter
    if filas:
        last_col = get_column_letter(len(COLUMNAS_DETALLE))
        ws.auto_filter.ref = f"A1:{last_col}{len(filas) + 1}"


# ════════════════════════════════════════════════════════════════════
# HOJA 2: RESUMEN EJECUTIVO
# ════════════════════════════════════════════════════════════════════

def construir_hoja_resumen(ws, filas):
    """Genera la hoja resumen con conteos."""
    from collections import Counter

    # Calcular metricas
    total = len(filas)

    # Por clasificacion
    por_clasif = Counter(f.get("clasificacion_operativa", "?") for f in filas)

    # Por antiguedad
    rangos = {"0-7 dias": 0, "8-30 dias": 0, "31-90 dias": 0, ">90 dias": 0, "Sin fecha": 0}
    for f in filas:
        dias = f.get("dias_abierta")
        if dias is None:
            rangos["Sin fecha"] += 1
        elif dias <= 7:
            rangos["0-7 dias"] += 1
        elif dias <= 30:
            rangos["8-30 dias"] += 1
        elif dias <= 90:
            rangos["31-90 dias"] += 1
        else:
            rangos[">90 dias"] += 1

    # Por operador
    por_operador = Counter(
        (f.get("operador") or "(sin operador)") for f in filas
    )

    # Por EECC
    por_eecc = Counter(
        (f.get("eecc") or "(sin EECC)") for f in filas
    )

    # Por tecnico (top 15)
    por_tecnico = Counter(
        (f.get("tecnico") or "(sin tecnico)") for f in filas
    )

    # Renderizar
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15

    fila = 1

    # Titulo principal
    cell = ws.cell(row=fila, column=1, value=f"REPORTE OTs INPRG MC - {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    cell.font = Font(name="Calibri", size=14, bold=True, color="1F4E78")
    fila += 2

    cell = ws.cell(row=fila, column=1, value="Total OTs INPRG MC en O_GESFO:")
    cell.font = Font(bold=True)
    cell = ws.cell(row=fila, column=2, value=total)
    cell.font = Font(bold=True, size=12)
    fila += 2

    # Por clasificacion
    cell = ws.cell(row=fila, column=1, value="Por clasificacion operativa")
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell = ws.cell(row=fila, column=2, value="Cantidad")
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell = ws.cell(row=fila, column=3, value="%")
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    fila += 1

    for clasif, cant in sorted(por_clasif.items(), key=lambda x: -x[1]):
        ws.cell(row=fila, column=1, value=clasif)
        ws.cell(row=fila, column=2, value=cant)
        ws.cell(row=fila, column=3, value=f"{cant/total*100:.1f}%" if total else "0%")
        fila += 1

    fila += 1

    # Por antiguedad
    cell = ws.cell(row=fila, column=1, value="Por antiguedad (dias abierta)")
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell = ws.cell(row=fila, column=2, value="Cantidad")
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell = ws.cell(row=fila, column=3, value="%")
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    fila += 1

    for rango, cant in rangos.items():
        ws.cell(row=fila, column=1, value=rango)
        ws.cell(row=fila, column=2, value=cant)
        ws.cell(row=fila, column=3, value=f"{cant/total*100:.1f}%" if total else "0%")
        fila += 1

    fila += 1

    # Por operador
    cell = ws.cell(row=fila, column=1, value="Por operador")
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell = ws.cell(row=fila, column=2, value="Cantidad")
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell = ws.cell(row=fila, column=3, value="%")
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    fila += 1

    for op, cant in sorted(por_operador.items(), key=lambda x: -x[1]):
        ws.cell(row=fila, column=1, value=op)
        ws.cell(row=fila, column=2, value=cant)
        ws.cell(row=fila, column=3, value=f"{cant/total*100:.1f}%" if total else "0%")
        fila += 1

    fila += 1

    # Por EECC
    cell = ws.cell(row=fila, column=1, value="Por EECC / Cuadrilla (top 15)")
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell = ws.cell(row=fila, column=2, value="Cantidad")
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell = ws.cell(row=fila, column=3, value="%")
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    fila += 1

    for eecc, cant in sorted(por_eecc.items(), key=lambda x: -x[1])[:15]:
        ws.cell(row=fila, column=1, value=eecc)
        ws.cell(row=fila, column=2, value=cant)
        ws.cell(row=fila, column=3, value=f"{cant/total*100:.1f}%" if total else "0%")
        fila += 1

    fila += 1

    # Por tecnico (top 15)
    cell = ws.cell(row=fila, column=1, value="Por tecnico asignado (top 15)")
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell = ws.cell(row=fila, column=2, value="Cantidad")
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell = ws.cell(row=fila, column=3, value="%")
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    fila += 1

    for tec, cant in sorted(por_tecnico.items(), key=lambda x: -x[1])[:15]:
        ws.cell(row=fila, column=1, value=tec)
        ws.cell(row=fila, column=2, value=cant)
        ws.cell(row=fila, column=3, value=f"{cant/total*100:.1f}%" if total else "0%")
        fila += 1


# ════════════════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════════════════

def main():
    print("="*70)
    print("REPORTE DE OTs INPRG MC - O_GESFO")
    print("="*70)

    # 1. Conectar a Postgres
    conn = obtener_conexion()
    if conn is None:
        print("[ERROR] No se pudo conectar a Postgres")
        return

    try:
        # 2. Ejecutar query
        print("\nConsultando Postgres...")
        from psycopg2.extras import RealDictCursor

        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(QUERY_INPRG)
            filas = [dict(r) for r in cur.fetchall()]

        print(f"OTs encontradas: {len(filas)}")

        if not filas:
            print("\n[INFO] No hay OTs en estado INPRG. Nada que reportar.")
            return

        # 3. Crear Excel
        print("\nGenerando Excel...")
        wb = Workbook()

        # Hoja 1: Detalle
        ws_detalle = wb.active
        ws_detalle.title = "OTs INPRG"
        construir_hoja_detalle(ws_detalle, filas)

        # Hoja 2: Resumen
        ws_resumen = wb.create_sheet(title="Resumen")
        construir_hoja_resumen(ws_resumen, filas)

        # 4. Guardar
        output_dir = Path("output")
        output_dir.mkdir(exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = output_dir / f"Reporte_INPRG_MC_{timestamp}.xlsx"

        wb.save(filename)

        print(f"\n[OK] Reporte generado: {filename}")
        print(f"     Total OTs:          {len(filas)}")
        print(f"     Hoja 1 (OTs INPRG): lista detallada")
        print(f"     Hoja 2 (Resumen):   conteos por antiguedad, operador, EECC, tecnico")

    except Exception as e:
        print(f"\n[ERROR] {type(e).__name__}: {e}")
        import traceback
        traceback.print_exc()

    finally:
        cerrar_conexion(conn)


if __name__ == "__main__":
    main()