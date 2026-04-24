"""
excel_exporter.py
----------------
Exporta registros OT a Excel multi-hoja con formato profesional.
Encapsula toda la logica de openpyxl (estilos, hojas, colores).

Hojas generadas:
    1. OTs O_GESFO 4213    -> todos los registros OT con todos los campos
    2. Worklogs            -> historial de avances (relacionado por wonum)
    3. Valores Válidos     -> catalogo de valores por spec
    4. Location vs Specs   -> specs tipicos por location (moda)
    5. Catalogo CINUM      -> cinums unicos con descripcion
    6. Leyenda             -> explicacion de colores
"""

import collections
import logging

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from domain.campos import (
    CAMPOS_OT, SPEC_CAMPOS, HARDCODEADOS, ANCHOS_COLUMNAS,
    WORKLOG_CAMPOS, ANCHOS_WORKLOG,
)
from exporters.base import Exporter


# ══════════════════════════════════════════════════════════════
# CONSTANTES DE ESTILO
# ══════════════════════════════════════════════════════════════

WHITE      = "FFFFFF"
COLOR_OT   = "1F4E79"  # azul oscuro   -> campos principales OT
COLOR_SPEC = "375623"  # verde oscuro  -> specs variables
COLOR_HARD = "7F3F00"  # naranja       -> specs hardcodeados
COLOR_WL   = "5B2C6F"  # morado oscuro -> worklogs
COLOR_ALT  = "F2F2F2"  # gris claro    -> filas alternas

_THIN   = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


# ══════════════════════════════════════════════════════════════
# HELPERS DE CELDAS
# ══════════════════════════════════════════════════════════════

def _hcell(ws, row, col, value, bg):
    """Celda de cabecera con fondo de color y texto blanco."""
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(bold=True, color=WHITE, name="Arial", size=9)
    c.fill      = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = _BORDER
    return c


def _dcell(ws, row, col, value, alt=False):
    """Celda de datos con fondo alterno opcional."""
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", size=9)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border    = _BORDER
    if alt:
        c.fill = PatternFill("solid", start_color=COLOR_ALT)
    return c


# ══════════════════════════════════════════════════════════════
# EXPORTER
# ══════════════════════════════════════════════════════════════

class ExcelExporter(Exporter):
    """Exporta a Excel multi-hoja con formato completo."""

    def __init__(self, output_file="tabla_maestra_cinum_gesfo.xlsx"):
        self.output_file = output_file

    def export(self, registros_ots, registros_worklogs=None):
        registros_worklogs = registros_worklogs or []

        wb = Workbook()
        self._hoja_ots(wb, registros_ots)
        self._hoja_worklogs(wb, registros_worklogs)
        self._hoja_valores_validos(wb, registros_ots)
        self._hoja_location_vs_specs(wb, registros_ots)
        self._hoja_catalogo_cinum(wb, registros_ots)
        self._hoja_leyenda(wb)
        wb.save(self.output_file)
        logging.info(
            f"Excel generado: {self.output_file} "
            f"({len(registros_ots)} OTs, {len(registros_worklogs)} worklogs)"
        )
        return f"Excel: {self.output_file}"

    # ── Hoja 1: OTs completas ──────────────────────────────────
    def _hoja_ots(self, wb, registros):
        ws = wb.active
        ws.title = "OTs O_GESFO 4213"
        cabeceras = CAMPOS_OT + SPEC_CAMPOS

        for col_idx, campo in enumerate(cabeceras, 1):
            if campo in HARDCODEADOS:
                bg = COLOR_HARD
            elif campo in SPEC_CAMPOS:
                bg = COLOR_SPEC
            else:
                bg = COLOR_OT
            _hcell(ws, 1, col_idx, campo, bg)

        for row_idx, reg in enumerate(registros, 2):
            alt = (row_idx % 2 == 0)
            for col_idx, campo in enumerate(cabeceras, 1):
                _dcell(ws, row_idx, col_idx, reg.get(campo, ""), alt)

        for col_idx, campo in enumerate(cabeceras, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = \
                ANCHOS_COLUMNAS.get(campo, 22)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cabeceras))}1"
        ws.row_dimensions[1].height = 35

    # ── Hoja 2: Worklogs ───────────────────────────────────────
    def _hoja_worklogs(self, wb, registros_worklogs):
        ws = wb.create_sheet("Worklogs")

        for col_idx, campo in enumerate(WORKLOG_CAMPOS, 1):
            _hcell(ws, 1, col_idx, campo, COLOR_WL)

        for row_idx, reg in enumerate(registros_worklogs, 2):
            alt = (row_idx % 2 == 0)
            for col_idx, campo in enumerate(WORKLOG_CAMPOS, 1):
                _dcell(ws, row_idx, col_idx, reg.get(campo, ""), alt)

        for col_idx, campo in enumerate(WORKLOG_CAMPOS, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = \
                ANCHOS_WORKLOG.get(campo, 22)

        ws.freeze_panes = "A2"
        if registros_worklogs:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(WORKLOG_CAMPOS))}1"
        ws.row_dimensions[1].height = 35

    # ── Hoja 3: Valores validos ────────────────────────────────
    def _hoja_valores_validos(self, wb, registros):
        ws = wb.create_sheet("Valores Válidos")
        _hcell(ws, 1, 1, "Campo",           COLOR_OT)
        _hcell(ws, 1, 2, "Tipo",            COLOR_OT)
        _hcell(ws, 1, 3, "Valores Válidos", COLOR_OT)

        valores_unicos = {}
        for reg in registros:
            for campo in SPEC_CAMPOS:
                val = reg.get(campo, "")
                if val:
                    valores_unicos.setdefault(campo, set()).add(str(val))

        for row, campo in enumerate(SPEC_CAMPOS, 2):
            valores = sorted(valores_unicos.get(campo, set()))
            tipo    = "Hardcodeado" if campo in HARDCODEADOS else "Variable"
            bg      = "FFE0CC" if campo in HARDCODEADOS else "D9E1F2"

            c1 = ws.cell(row=row, column=1, value=campo)
            c1.font   = Font(bold=True, name="Arial", size=9)
            c1.fill   = PatternFill("solid", start_color=bg)
            c1.border = _BORDER

            c2 = ws.cell(row=row, column=2, value=tipo)
            c2.font   = Font(name="Arial", size=9)
            c2.border = _BORDER

            c3 = ws.cell(row=row, column=3,
                         value=", ".join(valores) if valores else "—")
            c3.font      = Font(name="Arial", size=9)
            c3.alignment = Alignment(wrap_text=True)
            c3.border    = _BORDER

        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 90
        ws.row_dimensions[1].height = 20

    # ── Hoja 4: Location vs Specs ──────────────────────────────
    def _hoja_location_vs_specs(self, wb, registros):
        ws = wb.create_sheet("Location vs Specs")
        campos_loc = [
            "location", "nom_ubicacion", "cinum", "ci_description",
            "EECC_CUADRILLA_FO", "TIPO_CUADRILLA_FO", "OPERADOR_FO",
            "TIPO_OPERACION_FO", "TIPO_TRAMO", "COORDINADOR_RED_FO",
            "LIDER_DE_ZONA_FO", "RESPONSABLE_ZONA_NIVEL3_FO",
            "AREA_QUE_REPORTA_FO",
        ]
        campos_ot_base = {"location", "nom_ubicacion", "cinum", "ci_description"}

        for col_idx, campo in enumerate(campos_loc, 1):
            bg = COLOR_OT if campo in campos_ot_base else COLOR_SPEC
            _hcell(ws, 1, col_idx, campo, bg)

        loc_data = collections.defaultdict(lambda: collections.defaultdict(list))
        for reg in registros:
            loc = reg.get("location", "")
            if loc:
                for campo in campos_loc:
                    val = reg.get(campo, "")
                    if val:
                        loc_data[loc][campo].append(val)

        row = 2
        for loc, campos in sorted(loc_data.items()):
            alt = (row % 2 == 0)
            for col_idx, campo in enumerate(campos_loc, 1):
                valores = campos.get(campo, [])
                val = max(set(valores), key=valores.count) if valores else ""
                _dcell(ws, row, col_idx, val, alt)
            row += 1

        col_anchas = {
            "nom_ubicacion", "cinum", "ci_description",
            "EECC_CUADRILLA_FO", "COORDINADOR_RED_FO",
        }
        for col_idx, campo in enumerate(campos_loc, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = \
                28 if campo in col_anchas else 18

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(campos_loc))}1"
        ws.row_dimensions[1].height = 35

    # ── Hoja 5: Catalogo CINUM ─────────────────────────────────
    def _hoja_catalogo_cinum(self, wb, registros):
        ws = wb.create_sheet("Catalogo CINUM")
        _hcell(ws, 1, 1, "cinum",          COLOR_OT)
        _hcell(ws, 1, 2, "ci_description", COLOR_OT)

        cinums_unicos = {}
        for reg in registros:
            c = reg.get("cinum", "")
            if c and c not in cinums_unicos:
                cinums_unicos[c] = reg.get("ci_description", "")

        for row, (cinum, desc) in enumerate(sorted(cinums_unicos.items()), 2):
            alt = (row % 2 == 0)
            _dcell(ws, row, 1, cinum, alt)
            _dcell(ws, row, 2, desc, alt)

        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 55
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 25

    # ── Hoja 6: Leyenda ────────────────────────────────────────
    def _hoja_leyenda(self, wb):
        ws = wb.create_sheet("Leyenda")
        leyenda = [
            ("Color",        "Tipo",                   "Descripcion"),
            ("Azul oscuro",  "Campos OT principales",  "wonum, cinum, ci_description, location, status..."),
            ("Verde oscuro", "Specs variables",         "EECC_CUADRILLA_FO, TIPO_CAUSA, LIDER_DE_ZONA_FO..."),
            ("Naranja",      "Specs hardcodeados",      "OBSERV_CIERRE, COORDENADAS, PARADA_RELOJ..."),
            ("Morado oscuro","Hoja Worklogs",           "historial de avances de cada OT (relacion por wonum)"),
        ]
        for r_idx, (c1, c2, c3) in enumerate(leyenda, 1):
            ws.cell(row=r_idx, column=1, value=c1).font = \
                Font(bold=(r_idx == 1), name="Arial", size=10)
            ws.cell(row=r_idx, column=2, value=c2).font = Font(name="Arial", size=10)
            ws.cell(row=r_idx, column=3, value=c3).font = Font(name="Arial", size=10)

        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 60